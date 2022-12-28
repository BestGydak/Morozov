import csv

from datetime import datetime

import openpyxl.styles.numbers
from openpyxl.styles import Border, Font, Side
import openpyxl.styles.numbers
import openpyxl

import matplotlib
import matplotlib.pyplot as plt

import pdfkit
from jinja2 import Environment, FileSystemLoader

import tkinter





currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
class Report:
    def __init__(self, font, header_font, border,
                 _salary_dynamic, _count_dynamic,
                 _selected_salary_dynamic, _selected_count_dynamic,
                 _top_cities, _top_cities_count, _vacancy_name):
        self._wb = openpyxl.Workbook()
        self._header_font = header_font
        self._font = font
        self._border = border
        self._salary_dynamic = _salary_dynamic
        self._count_dynamic = _count_dynamic
        self._selected_salary_dynamic = _selected_salary_dynamic
        self._selected_count_dynamic = _selected_count_dynamic
        self._top_cities = _top_cities
        self._top_cities_count = _top_cities_count
        self._vacancy_name = _vacancy_name

    def generate_charts(self):
        matplotlib.use('TkAgg')
        plt.rc('font', size=8)
        plt.rc('axes', titlesize=13)

        figure = plt.figure()

        self.add_salary_dynamic_chart(figure)
        self.add_count_dynamic_chart(figure)
        self.add_top_cities_chart(figure)
        self.add_top_cities_count_chart(figure)

        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        plt.show()
        figure.savefig('graph.png')

    def add_salary_dynamic_chart(self, figure):
        subplot = figure.add_subplot(2, 2, 1)
        x = list(self._salary_dynamic)
        y1 = [self._salary_dynamic[k] for k in x]
        y2 = [self._selected_salary_dynamic[k] for k in x]
        subplot.bar([i - 0.2 for i in x], y1, label='средняя з/п', width=0.4)
        subplot.bar([i + 0.2 for i in x], y2, label=f'з/п {self._vacancy_name}', width=0.4)
        subplot.set_xticks(x, fontsize=8)
        subplot.tick_params(axis='x', labelrotation=90)
        subplot.grid(axis='y')
        subplot.set(title='Уровень зарплат по годам')

    def add_count_dynamic_chart(self, figure):
        subplot = figure.add_subplot(2, 2, 2)
        x = list(self._count_dynamic)
        y1 = [self._count_dynamic[k] for k in x]
        y2 = [self._selected_count_dynamic[k] for k in x]
        subplot.bar([i - 0.2 for i in x], y1, label='Количество вакансий', width=0.4)
        subplot.bar([i + 0.2 for i in x], y2, label=f'Количество вакансий {self._vacancy_name}', width=0.4)
        subplot.set_xticks(x, fontsize=8)
        subplot.tick_params(axis='x', labelrotation=90)
        subplot.grid(axis='y')
        subplot.set(title='Количество вакансий по годам')

    def add_top_cities_chart(self, figure):
        subplot = figure.add_subplot(2, 2, 3)
        x = list(self._top_cities)[:10]
        y = [self._top_cities[k] for k in x]
        x = [i if ' ' not in i and '-' not in i else i.replace(' ', '\n').replace('-', '-\n') for i in x]
        subplot.barh(x, y)
        subplot.tick_params(axis='y', labelsize=6)
        subplot.grid(axis='x')
        subplot.set(title='Уровень зарплат по городам')

    def add_top_cities_count_chart(self, figure):
        subplot = figure.add_subplot(2, 2, 4)
        x = list(self._top_cities_count)[:10][::-1]
        y = [self._top_cities_count[k] for k in x]
        x.append('Другие')
        y.append(1 - sum(y))
        subplot.pie(y, labels=x)
        subplot.set(title='Доля вакансий по городам')

    def generate_excel(self):
        ws = self._wb.active
        ws.title = 'Статистика по годам'
        header = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self._vacancy_name}', 'Количество вакансий',
                  f'Количество вакансий - {self._vacancy_name}']
        ws.append(header)
        for year in self._salary_dynamic:
            row = [year]
            if year in self._salary_dynamic:
                row.append(self._salary_dynamic[year])
            if year in self._count_dynamic:
                row.append(self._count_dynamic[year])
            if year in self._selected_salary_dynamic:
                row.append( self._selected_salary_dynamic[year])
            if year in self._selected_count_dynamic:
                row.append( self._selected_count_dynamic[year])
            if year in self._top_cities:
                row.append(self._top_cities[year])
            if year in self._top_cities_count:
                row.append(self._top_cities_count[year])
            ws.append(row)
        self._formating_sheet(len(header), len(salary_dynamic), header, ws)

        ws = self._wb.create_sheet('Статистика по городам')
        header = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        ws.append(header)
        for city, _ in zip(self._top_cities_count, range(10)):
            row = [city]
            row.append(self._top_cities[city])
            row.append('')
            row.append(city)
            row.append(self._top_cities_count[city])
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=11, min_col=5):
            for cell in row:
                cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]
        self._formating_sheet(len(header), 10, header, ws)
        self._wb.save('report.xlsx')

    def _calculate_column_width(self, column_values):
        max_width = len(str(max(column_values, key = lambda x: len(str(x)))))
        return  max_width + 2

    def _formating_sheet(self, row_len, column_len, header, sheet):
        A_ord = ord('A')
        for column_ord in range(row_len):
            column = chr(column_ord + A_ord)
            sheet.column_dimensions[column].width = self._calculate_column_width([v.value for v in sheet[column]])
        for row in sheet.iter_rows(max_row=1, max_col=row_len):
            for cell in row:
                cell.font = self._header_font
                cell.border = self._border
        for row in sheet.iter_rows(max_row=column_len+1, min_row=2, max_col=row_len):
            for cell in row:
                cell.font = self._font
                cell.border = self._border

    def generate_pdf(self):

        self.generate_charts()
        self.generate_excel()
        header2 = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        stats = [
            {x: x for x in self._salary_dynamic},
            self._salary_dynamic,
            self._selected_salary_dynamic,
            self._count_dynamic,
            self._selected_count_dynamic,
            {y: y for y in top_cities_count},
            top_cities_count,
            top_cities]
        header = ['\tГод\t', 'Средняя зарплата', f'Средняя зарплата - {self._vacancy_name}', 'Количество вакансий',
                  f'Количество вакансий - {self._vacancy_name}']
        pdf_template = template.render({'header': header, 'image_file': 'graph.png', 'round': round,
                                        'stats': stats, 'vacancy_name': vacancy_name, 'header2': header2})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config)



class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name,
                 published_at):
        self.name = name
        self.salary_from = float(salary_from)
        self.salary_to = float(salary_to)
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z")

    def get_salary_in_rub(self):
        return currency_to_rub[self.salary_currency] * (self.salary_to + self.salary_from) / 2


def csv_reader(file_name):
    vacancies = []
    titles = []
    with open(file_name, 'r', encoding='utf-8-sig') as csvfile:
        reader = csv.reader(csvfile)
        is_first = True
        for line in reader:
            if is_first:
                titles = line
                is_first = False
                continue
            if len(line) == len(titles) and "" not in line:
                vacancies.append(line)
    return vacancies, titles


def csv_filer(vacancies_list, titles):
    vacancies_objects = []
    for info in vacancies_list:
        new_vacancy = Vacancy(info[titles.index('name')],
                              info[titles.index('salary_from')],
                              info[titles.index('salary_to')],
                              info[titles.index('salary_currency')],
                              info[titles.index('area_name')],
                              info[titles.index('published_at')]
                              )
        vacancies_objects.append(new_vacancy)
    return vacancies_objects


def sort_salaries_by_year(vacancies):
    vacancies_by_year = dict()
    for vacancy in vacancies:
        year = vacancy.published_at.year
        if vacancies_by_year.get(year) is None:
            vacancies_by_year[year] = [vacancy]
        else:
            vacancies_by_year[year].append(vacancy)
    temp_list = [[key, vacancies_by_year[key]] for key in vacancies_by_year.keys()]
    temp_list.sort(key=lambda x: x[0])
    result_dict = dict()
    for info in temp_list:
        result_dict[info[0]] = info[1]
    return result_dict


def get_salary_dynamic(vacancies_by_year):
    salaries_dynamic = dict()
    for year in sorted(vacancies_by_year.keys()):
        if len(vacancies_by_year[year]) == 0:
            salaries_dynamic[year] = 0
            continue
        avrg = int(sum(v.get_salary_in_rub() for v in vacancies_by_year[year]) / len(vacancies_by_year[year]))
        salaries_dynamic[year] = avrg
    return salaries_dynamic


def get_count_dynamic(vacancies_by_year):
    count_dynamic = dict()
    for year in sorted(vacancies_by_year.keys()):
        count_dynamic[year] = len(vacancies_by_year[year])
    return count_dynamic

def get_big_enough_cities(vacancies):
    cities = dict()
    for vacancy in vacancies:
        if not (vacancy.area_name in cities.keys()):
            cities[vacancy.area_name] = 1
        else:
            cities[vacancy.area_name] += 1
    vacancies_count = len(vacancies)
    allowed_cities = set()
    for city in cities.keys():
        if cities[city] / vacancies_count >= 0.01:
            allowed_cities.add(city)
    return allowed_cities


def get_top_cities(vacancies, allowed_cities):
    cities = dict()
    for vacancy in vacancies:
        if not(vacancy.area_name in allowed_cities):
            continue
        if not (vacancy.area_name in cities.keys()):
            cities[vacancy.area_name] = [vacancy.get_salary_in_rub(), 1]
        else:
            cities[vacancy.area_name][0] += vacancy.get_salary_in_rub()
            cities[vacancy.area_name][1] += 1
    temp_list = [[key, int(cities[key][0] / cities[key][1])] for key in cities.keys()]
    temp_list.sort(key=lambda x: x[1], reverse=True)
    result_dict = dict()
    for info in temp_list:
        result_dict[info[0]] = info[1]
    return result_dict


def get_top_cities_count(vacancies, allowed_cities):
    cities = dict()
    for vacancy in vacancies:
        if not(vacancy.area_name in allowed_cities):
            continue
        if not (vacancy.area_name in cities.keys()):
            cities[vacancy.area_name] = 1
        else:
            cities[vacancy.area_name] += 1
    vacancies_count = len(vacancies)

    temp_list = [[key, round(cities[key] / vacancies_count, 4)] for key in cities.keys()]
    temp_list.sort(key=lambda x: x[1], reverse=True)
    temp_list = list(filter(lambda x: x[1] >= 0.01, temp_list))

    result_dict = dict()
    for info in temp_list:
        result_dict[info[0]] = info[1]
    return result_dict


file_name = input("Введите название файла: ")
vacancy_name = input("Введите название профессии: ")
vacancies_info, titles = csv_reader(file_name)
all_vacancies = csv_filer(vacancies_info, titles)
chosen_vacancies = list(filter(lambda x: vacancy_name in x.name, all_vacancies))
all_vacancies_by_year = sort_salaries_by_year(all_vacancies)
chosen_vacancies_by_year = dict()
for key in all_vacancies_by_year.keys():
    chosen_vacancies_by_year[key] = list(filter(lambda x: vacancy_name in x.name, all_vacancies_by_year[key]))
allowed_cities = get_big_enough_cities(all_vacancies)

salary_dynamic = get_salary_dynamic(all_vacancies_by_year)
count_dynamic = get_count_dynamic(all_vacancies_by_year)
selected_salary_dynamic = get_salary_dynamic(chosen_vacancies_by_year)
selected_count_dynamic = get_count_dynamic(chosen_vacancies_by_year)
top_cities = get_top_cities(all_vacancies, allowed_cities)
top_cities_count = get_top_cities_count(all_vacancies, allowed_cities)


font = Font(name='Calibri', size=11)
header_font = Font(name='Calibri', size=11, bold=True)
border = Border(right=Side(color='000000', border_style='thin'), bottom=Side(color='000000', border_style='thin'))
report = Report(font, header_font, border, salary_dynamic, count_dynamic, selected_salary_dynamic,
                selected_count_dynamic, top_cities, top_cities_count, vacancy_name)

report.generate_pdf()
print("Динамика уровня зарплат по годам:",salary_dynamic)
print("Динамика количества вакансий по годам:", count_dynamic)
print("Динамика уровня зарплат по годам для выбранной профессии:", selected_salary_dynamic)
print("Динамика количества вакансий по годам для выбранной профессии:", selected_count_dynamic)
print("Уровень зарплат по городам (в порядке убывания):", {k: top_cities[k] for i, k in zip(range(10), top_cities)})
print("Доля вакансий по городам (в порядке убывания):", {k: top_cities_count[k] for i, k in zip(range(10),
                                                                                                top_cities_count)})
